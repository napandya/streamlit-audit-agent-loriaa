"""
Export functionality for audit data
"""
import os
import tempfile

import streamlit as st
import pandas as pd
from typing import List
from datetime import datetime
import io

from models.unit import Unit, RecurringTransaction, AuditFinding
from utils.helpers import format_currency
from storage.audit_log import AuditLog


def render_export_panel(
    units: List[Unit],
    transactions: List[RecurringTransaction],
    findings: List[AuditFinding],
    audit_log: AuditLog
):
    """
    Render export panel with download options
    """
    st.header("📤 Export Data")
    
    st.write("Download audit data and findings in various formats.")
    
    # Export format selector
    export_format = st.radio(
        "Select export format:",
        options=["Excel (recommended)", "CSV"],
        help="Choose the format for exporting data"
    )
    
    # Export options
    st.subheader("Select data to export:")
    
    col1, col2 = st.columns(2)
    
    with col1:
        export_findings = st.checkbox("Audit Findings", value=True)
        export_units = st.checkbox("Unit Summary", value=True)
    
    with col2:
        export_transactions = st.checkbox("All Transactions", value=False)
        export_summary = st.checkbox("Executive Summary", value=True)
    
    # User name for audit trail
    exporter_name = st.text_input(
        "Your name (for audit trail):",
        placeholder="Enter your name"
    )
    
    # Export button
    if st.button("📥 Generate Export", type="primary", use_container_width=True):
        if not exporter_name:
            st.error("Please enter your name for the audit trail.")
        else:
            with st.spinner("Generating export..."):
                if export_format == "Excel (recommended)":
                    export_data = generate_excel_export(
                        units, transactions, findings,
                        export_findings, export_units, export_transactions, export_summary
                    )
                    
                    filename = f"audit_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    
                else:  # CSV
                    export_data = generate_csv_export(findings)
                    filename = f"audit_findings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                    mime_type = "text/csv"
                
                # Log export action
                audit_log.log_export(
                    export_type=export_format,
                    user=exporter_name,
                    record_count=len(findings)
                )
                
                # Download button
                st.download_button(
                    label=f"💾 Download {filename}",
                    data=export_data,
                    file_name=filename,
                    mime=mime_type,
                    use_container_width=True
                )
                
                st.success("✅ Export generated successfully!")


def generate_excel_export(
    units: List[Unit],
    transactions: List[RecurringTransaction],
    findings: List[AuditFinding],
    include_findings: bool,
    include_units: bool,
    include_transactions: bool,
    include_summary: bool
) -> bytes:
    """Generate Excel file with multiple sheets"""
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # Executive Summary Sheet
        if include_summary:
            summary_data = generate_summary_data(units, transactions, findings)
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
        
        # Findings Sheet
        if include_findings and findings:
            findings_df = generate_findings_dataframe(findings)
            findings_df.to_excel(writer, sheet_name='Audit Findings', index=False)
        
        # Units Sheet
        if include_units and units:
            units_df = generate_units_dataframe(units, transactions, findings)
            units_df.to_excel(writer, sheet_name='Unit Summary', index=False)
        
        # Transactions Sheet
        if include_transactions and transactions:
            transactions_df = generate_transactions_dataframe(transactions)
            transactions_df.to_excel(writer, sheet_name='All Transactions', index=False)
    
    output.seek(0)
    return output.getvalue()


def generate_csv_export(findings: List[AuditFinding]) -> bytes:
    """Generate CSV file with findings"""
    df = generate_findings_dataframe(findings)
    
    output = io.StringIO()
    df.to_csv(output, index=False)
    output.seek(0)
    
    return output.getvalue().encode('utf-8')


def generate_summary_data(
    units: List[Unit],
    transactions: List[RecurringTransaction],
    findings: List[AuditFinding]
) -> List[dict]:
    """Generate executive summary data"""
    
    total_revenue = sum(t.amount for t in transactions if t.category in ['rent', 'fee'])
    total_concessions = sum(abs(t.amount) for t in transactions if t.category == 'concession')
    net_revenue = total_revenue - total_concessions
    
    summary = [
        {'Metric': 'Total Units', 'Value': len(units)},
        {'Metric': 'Total Revenue', 'Value': format_currency(total_revenue)},
        {'Metric': 'Total Concessions', 'Value': format_currency(total_concessions)},
        {'Metric': 'Net Revenue', 'Value': format_currency(net_revenue)},
        {'Metric': '', 'Value': ''},
        {'Metric': 'Total Findings', 'Value': len(findings)},
        {'Metric': 'Critical Findings', 'Value': len([f for f in findings if f.severity == 'Critical'])},
        {'Metric': 'High Findings', 'Value': len([f for f in findings if f.severity == 'High'])},
        {'Metric': 'Medium Findings', 'Value': len([f for f in findings if f.severity == 'Medium'])},
        {'Metric': 'Low Findings', 'Value': len([f for f in findings if f.severity == 'Low'])},
        {'Metric': '', 'Value': ''},
        {'Metric': 'Open Findings', 'Value': len([f for f in findings if f.status == 'Open'])},
        {'Metric': 'Reviewed Findings', 'Value': len([f for f in findings if f.status == 'Reviewed'])},
        {'Metric': 'Overridden Findings', 'Value': len([f for f in findings if f.status == 'Overridden'])},
        {'Metric': 'Closed Findings', 'Value': len([f for f in findings if f.status == 'Closed'])},
    ]
    
    return summary


def generate_findings_dataframe(findings: List[AuditFinding]) -> pd.DataFrame:
    """Generate findings dataframe for export"""
    
    from engine.explainability import ExplainabilityEngine
    
    data = []
    for finding in findings:
        explanation = ExplainabilityEngine.explain(finding)
        
        data.append({
            'Finding ID': finding.finding_id,
            'Unit Number': finding.unit_number,
            'Rule': finding.rule_name,
            'Severity': finding.severity,
            'Month': finding.month.strftime('%b %Y') if finding.month else 'N/A',
            'Delta': finding.delta if finding.delta else 0,
            'Explanation': explanation,
            'Status': finding.status,
            'Notes': finding.notes,
            'Reviewed By': finding.reviewed_by or '',
            'Reviewed At': finding.reviewed_at.strftime('%Y-%m-%d') if finding.reviewed_at else '',
        })
    
    return pd.DataFrame(data)


def generate_units_dataframe(
    units: List[Unit],
    transactions: List[RecurringTransaction],
    findings: List[AuditFinding]
) -> pd.DataFrame:
    """Generate units summary dataframe"""
    
    from collections import defaultdict
    
    # Aggregate by unit
    unit_totals = defaultdict(lambda: {'rent': 0, 'concessions': 0, 'fees': 0})
    
    for txn in transactions:
        if txn.category == 'rent':
            unit_totals[txn.unit_id]['rent'] += txn.amount
        elif txn.category == 'concession':
            unit_totals[txn.unit_id]['concessions'] += abs(txn.amount)
        elif txn.category == 'fee':
            unit_totals[txn.unit_id]['fees'] += txn.amount
    
    # Count findings per unit
    unit_findings = defaultdict(int)
    for finding in findings:
        unit_findings[finding.unit_id] += 1
    
    data = []
    for unit in units:
        totals = unit_totals[unit.unit_id]
        
        data.append({
            'Unit Number': unit.unit_number,
            'Resident Name': unit.resident_name or 'Vacant',
            'Employee Unit': 'Yes' if unit.is_employee_unit else 'No',
            'Base Rent': unit.base_rent or 0,
            'Total Rent': totals['rent'],
            'Total Concessions': totals['concessions'],
            'Total Fees': totals['fees'],
            'Net Revenue': totals['rent'] + totals['fees'] - totals['concessions'],
            'Findings Count': unit_findings[unit.unit_id],
            'Lease Start': unit.lease_start.strftime('%Y-%m-%d') if unit.lease_start else '',
            'Lease End': unit.lease_end.strftime('%Y-%m-%d') if unit.lease_end else '',
        })
    
    return pd.DataFrame(data)


def generate_transactions_dataframe(transactions: List[RecurringTransaction]) -> pd.DataFrame:
    """Generate transactions dataframe"""
    
    data = []
    for txn in transactions:
        data.append({
            'Transaction ID': txn.transaction_id,
            'Unit Number': txn.unit_number,
            'Month': txn.month.strftime('%b %Y') if txn.month else 'N/A',
            'Category': txn.category.title(),
            'Subcategory': txn.subcategory or '',
            'Description': txn.description,
            'Amount': txn.amount,
            'Source': txn.source,
        })
    
    return pd.DataFrame(data)


# ---------------------------------------------------------------------------
# v4 Forensic Audit Excel Export
# ---------------------------------------------------------------------------

def export_audit_workbook(results: dict, output_path: str) -> str:
    """
    Generate a color-coded multi-sheet Excel workbook from resman_results.

    Sheets:
      John's Flags            — John's 9 concession rule violations
      Daniel's Flags          — Daniel's 2-stage revenue integrity flags
      Fee Schedule Violations — per-property fee amount violations
      Manager Overrides       — manager edit/reversal leaderboard
      Override Detail Log     — raw override event log
      Exposure Summary        — by_property, by_rule, by_risk rollups

    Risk color coding: CRITICAL=#FF4B4B, HIGH=#FFA500, MEDIUM=#FFD700

    Parameters
    ----------
    results:
        Dict returned by ``engine.resman_rules.run_resman_audit``.
    output_path:
        Destination ``.xlsx`` file path.

    Returns
    -------
    str: The path written (same as *output_path*).
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install with: pip install openpyxl"
        )

    RISK_FILLS = {
        "CRITICAL": PatternFill("solid", fgColor="FF4B4B"),
        "HIGH": PatternFill("solid", fgColor="FFA500"),
        "MEDIUM": PatternFill("solid", fgColor="FFD700"),
    }
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

    johns_flags = results.get("johns_flags", pd.DataFrame())
    daniels_flags = results.get("daniels_flags", pd.DataFrame())
    fee_flags = results.get("fee_flags", pd.DataFrame())
    manager_ranking = results.get("manager_ranking", pd.DataFrame())
    override_log = results.get("override_log", pd.DataFrame())
    exposure = results.get("exposure", {})

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if not johns_flags.empty:
            johns_flags.to_excel(writer, sheet_name="John's Flags", index=False)
        if not daniels_flags.empty:
            daniels_flags.to_excel(writer, sheet_name="Daniel's Flags", index=False)
        if not fee_flags.empty:
            fee_flags.to_excel(writer, sheet_name="Fee Schedule Violations", index=False)
        if not manager_ranking.empty:
            manager_ranking.to_excel(writer, sheet_name="Manager Overrides", index=False)
        if not override_log.empty:
            override_log.to_excel(writer, sheet_name="Override Detail Log", index=False)
        for key, label in [
            ("by_property", "Exposure by Property"),
            ("by_rule", "Exposure by Rule"),
            ("by_risk", "Exposure by Risk"),
        ]:
            df_exp = exposure.get(key, pd.DataFrame())
            if not df_exp.empty:
                df_exp.to_excel(writer, sheet_name=label, index=False)

    # Post-process: apply colour coding and header formatting
    wb = load_workbook(output_path)

    flag_sheets = ["John's Flags", "Daniel's Flags", "Fee Schedule Violations"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Header row formatting
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

        # Risk column color coding (flag sheets only)
        if sheet_name in flag_sheets:
            risk_col_idx = None
            for col_idx, cell in enumerate(ws[1], start=1):
                if str(cell.value).strip() == "Risk_Level":
                    risk_col_idx = col_idx
                    break

            if risk_col_idx is not None:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    risk_cell = row[risk_col_idx - 1]
                    fill = RISK_FILLS.get(str(risk_cell.value or ""), None)
                    if fill:
                        risk_cell.fill = fill

        # Auto-fit column widths
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
            for row_idx in range(2, min(ws.max_row + 1, 200)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, min(len(str(val)), 60))
            ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

    wb.save(output_path)
    return output_path


def render_excel_download_button(results: dict) -> None:
    """
    Streamlit helper — generates an Excel audit workbook and renders a
    ``st.download_button`` for the ``.xlsx`` file.

    Parameters
    ----------
    results:
        Dict returned by ``engine.resman_rules.run_resman_audit``.
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    fname = f"LNJ_Audit_{ts}.xlsx"

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = os.path.join(tmp_dir, fname)
        try:
            export_audit_workbook(results, tmp_path)
            with open(tmp_path, "rb") as f:
                xlsx_bytes = f.read()
        except Exception as exc:
            st.error(f"Excel export failed: {exc}")
            return

    st.download_button(
        label="⬇️ Download Audit Workbook (.xlsx)",
        data=xlsx_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
